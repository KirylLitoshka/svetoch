import io
from datetime import datetime
import json

from sqlalchemy import select, literal_column, func, desc, and_
from openpyxl import load_workbook

from views import ListView, DetailView, BaseView
from utils import pretty_json, DATE_FORMAT
from table import get_table_data
from warmth.models import *
from warmth.calculations import *
from warmth.queries import *
from warmth.reports import *


class SubsystemDetailView(DetailView):
    model = subsystem

    async def get(self):
        async with self.request.app["db"].connect() as conn:
            cursor = await conn.execute(select(self.model))
            result = dict(cursor.fetchone())
            return web.json_response(
                {"success": True, "item": result}, dumps=pretty_json
            )

    async def patch(self):
        current_data = await self.request.json()
        async with self.request.app["db"].begin() as conn:
            await conn.execute(
                self.model.update()
                .where(self.model.c.id == current_data["id"])
                .values(**current_data)
            )
        self.request.app["subsystem"] = current_data
        return web.json_response({"success": True}, dumps=pretty_json)


class BanksListView(ListView):
    model = banks


class BankDetailView(DetailView):
    model = banks


class RatesListView(ListView):
    model = rates

    async def get(self):
        async with self.request.app["db"].connect() as conn:
            limited_history = (
                select(rates_history.c.id)
                .where(rates_history.c.rate_id == literal_column("rates.id"))
                .order_by(desc(rates_history.c.id))
                .limit(1)
            )
            smtm = select(
                rates, func.row_to_json(rates_history.table_valued()).label("history")
            ).select_from(
                rates.join(
                    rates_history, rates_history.c.id == limited_history, isouter=True
                )
            )
            cursor = await conn.execute(smtm)
            result = [dict(row) for row in cursor.fetchall()]
            return web.json_response(
                {"success": True, "items": result}, dumps=pretty_json
            )


class RatesDetailView(DetailView):
    model = rates


class RatesHistoryListView(ListView):
    model = rates_history

    async def get(self):
        rate_id = int(self.request.match_info["rate_id"])
        async with self.request.app["db"].connect() as conn:
            cursor = await conn.execute(
                select(self.model)
                .where(self.model.c.rate_id == rate_id)
                .order_by(desc(self.model.c.id))
            )
            result = [dict(row) for row in cursor.fetchall()]
            return web.json_response({"success": True, "items": result})


class RateHistoryDetailView(DetailView):
    model = rates_history


class WorkshopsListView(ListView):
    model = workshops

    async def get(self):
        async with self.request.app["db"].connect() as conn:
            cursor = await conn.execute(
                select(
                    self.model.c.id,
                    self.model.c.title,
                    self.model.c.is_currency_coefficient_applied,
                    func.row_to_json(workshops_groups.table_valued()).label("group"),
                )
                .select_from(self.model.join(workshops_groups, isouter=True))
                .order_by(self.model.c.id)
            )
            result = [dict(row) for row in cursor.fetchall()]
            return web.json_response(
                {"success": True, "items": result}, dumps=pretty_json
            )


class WorkshopDetailView(DetailView):
    model = workshops


class CurrencyCoefficientsListView(ListView):
    model = currency_coefficients

    async def get(self):
        async with self.request.app["db"].connect() as conn:
            cursor = await conn.execute(
                select(self.model).order_by(
                    desc(self.model.c.year), desc(self.model.c.month)
                )
            )
            result = [dict(row) for row in cursor.fetchall()]
            return web.json_response(
                {"success": True, "items": result}, dumps=pretty_json
            )


class CurrencyCoefficientDetailView(DetailView):
    model = currency_coefficients


class ReconciliationCodesListView(ListView):
    model = reconciliation_codes


class ReconciliationCodePayments(ListView):
    model = payments

    async def get(self):
        code_id = int(self.request.match_info["id"])
        app_info = self.request.app["subsystem"]
        async with self.request.app["db"].connect() as conn:
            cursor = await conn.execute(
                select(objects.c.code, objects.c.title, self.model)
                .select_from(self.model.join(objects))
                .where(
                    and_(
                        objects.c.reconciliation_code_id == code_id,
                        self.model.c.operation_year == app_info["year"],
                        self.model.c.operation_month == app_info["month"],
                    )
                )
            )
            result = [dict(row) for row in cursor.fetchall()]
            if not result:
                return web.json_response(
                    {
                        "success": False,
                        "reason": "Начислений по указанному коду не найдено",
                    }
                )
            return web.json_response(
                {"success": True, "items": result}, dumps=pretty_json
            )


class ReconciliationCodeDetailView(DetailView):
    model = reconciliation_codes


class ObjectsListView(ListView):
    model = objects

    async def get(self):
        async with self.request.app["db"].connect() as conn:
            cursor = await conn.execute(
                select(
                    self.model,
                    func.row_to_json(rates.table_valued()).label("rate"),
                    func.row_to_json(workshops.table_valued()).label("workshop"),
                    func.row_to_json(reconciliation_codes.table_valued()).label(
                        "reconciliation_code"
                    ),
                )
                .select_from(
                    self.model.join(rates)
                    .join(workshops)
                    .join(reconciliation_codes, isouter=True)
                )
                .order_by(self.model.c.code)
            )
            result = [dict(row) for row in cursor.fetchall()]
            return web.json_response({"success": True, "items": result})

    async def post(self):
        post_data = await self.request.json()
        code_id = (
            post_data["reconciliation_code"]["id"]
            if post_data["reconciliation_code"] is not None
            else None
        )
        async with self.request.app["db"].begin() as conn:
            cursor = await conn.execute(
                self.model.insert()
                .returning(self.model.c.id)
                .values(
                    title=post_data["title"],
                    code=post_data["code"],
                    rate_id=(
                        post_data["rate"]["id"]
                        if post_data["rate"] is not None
                        else None
                    ),
                    workshop_id=(
                        post_data["workshop"]["id"]
                        if post_data["workshop"] is not None
                        else None
                    ),
                    reconciliation_code_id=code_id,
                    is_closed=post_data["is_closed"],
                    vat=post_data["vat"],
                    is_meter_unavailable=post_data["is_meter_unavailable"],
                    is_heating_available=post_data["is_heating_available"],
                    heating_load=post_data["heating_load"],
                    is_water_heating_available=post_data["is_water_heating_available"],
                    water_heating_load=post_data["water_heating_load"],
                )
            )
            post_data["id"] = dict(cursor.fetchone()).get("id")
            return web.json_response({"success": True, "item": post_data})


class ObjectDetailView(DetailView):
    model = objects

    async def patch(self):
        post_data = await self.request.json()
        object_id = int(self.request.match_info["id"])
        code_id = (
            post_data["reconciliation_code"]["id"]
            if post_data["reconciliation_code"] is not None
            else None
        )
        async with self.request.app["db"].begin() as conn:
            await conn.execute(
                self.model.update()
                .where(self.model.c.id == object_id)
                .values(
                    title=post_data["title"],
                    code=post_data["code"],
                    rate_id=(
                        post_data["rate"]["id"]
                        if post_data["rate"] is not None
                        else None
                    ),
                    workshop_id=(
                        post_data["workshop"]["id"]
                        if post_data["workshop"] is not None
                        else None
                    ),
                    reconciliation_code_id=code_id,
                    is_closed=post_data["is_closed"],
                    vat=post_data["vat"],
                    is_meter_unavailable=post_data["is_meter_unavailable"],
                    is_heating_available=post_data["is_heating_available"],
                    heating_load=post_data["heating_load"],
                    is_water_heating_available=post_data["is_water_heating_available"],
                    water_heating_load=post_data["water_heating_load"],
                )
            )
            return web.json_response({"success": True}, dumps=pretty_json)


class RentersListView(ListView):
    model = renters

    async def get(self):
        async with self.request.app["db"].connect() as conn:
            cursor = await conn.execute(
                select(self.model, func.row_to_json(banks.table_valued()).label("bank"))
                .select_from(self.model.join(banks, isouter=True))
                .order_by(self.model.c.id)
            )
            result = [dict(row) for row in cursor.fetchall()]
            return web.json_response(
                {"success": True, "items": result}, dumps=pretty_json
            )

    async def post(self):
        post_data = await self.request.json()
        bank = post_data.pop("bank", None)
        contract_date = post_data.get("contract_date")
        if contract_date:
            post_data["contract_date"] = datetime.strptime(contract_date, DATE_FORMAT)
        async with self.request.app["db"].begin() as conn:
            cursor = await conn.execute(
                self.model.insert().returning(literal_column("*")).values(**post_data)
            )
            result = dict(cursor.fetchone())
            result["bank"] = bank
            return web.json_response(
                {"success": True, "item": result}, dumps=pretty_json
            )


class RenterDetailView(DetailView):
    model = renters

    async def patch(self):
        renter_id = int(self.request.match_info["id"])
        post_data = await self.request.json()
        if post_data["contract_date"]:
            post_data["contract_date"] = datetime.strptime(
                post_data.pop("contract_date"), DATE_FORMAT
            )
        async with self.request.app["db"].begin() as conn:
            await conn.execute(
                self.model.update()
                .where(self.model.c.id == renter_id)
                .values(
                    **{
                        k: v
                        for k, v in post_data.items()
                        if k in self.model.columns.keys()
                    }
                )
            )
            return web.json_response({"success": True}, dumps=pretty_json)


class RentersObjectsListView(ListView):
    model = renters_objects

    async def get(self):
        renter_id = int(self.request.match_info["renter_id"])
        async with self.request.app["db"].connect() as conn:
            cursor = await conn.execute(
                self.model.select().where(self.model.c.renter_id == renter_id)
            )
            result = [dict(row) for row in cursor.fetchall()]
            return web.json_response(
                {"success": True, "items": result}, dumps=pretty_json
            )

    async def post(self):
        post_data = await self.request.json()
        async with self.request.app["db"].begin() as conn:
            cursor = await conn.execute(
                self.model.insert().returning(literal_column("*")).values(**post_data)
            )
            result = dict(cursor.fetchone())
            return web.json_response(
                {"success": True, "item": result}, dumps=pretty_json
            )

    async def delete(self):
        renter_id = int(self.request.match_info["renter_id"])
        obj_id = int(self.request.match_info["id"])
        async with self.request.app["db"].begin() as conn:
            await conn.execute(
                self.model.delete().where(
                    and_(
                        self.model.c.renter_id == renter_id,
                        self.model.c.object_id == obj_id,
                    )
                )
            )
            return web.json_response({"success": True}, dumps=pretty_json)


class PaymentsUploadView(BaseView):
    model = payments

    async def post(self):
        post_data = await self.request.post()
        table_data = await get_table_data(
            post_data["file"].file, self.request.app["app_name"]
        )
        required_codes = {row["KO"] for row in table_data}
        app_info = self.request.app["subsystem"]
        async with self.request.app["db"].begin() as conn:
            cursor = await conn.execute(
                objects.select().where(objects.c.code.in_(required_codes))
            )
            selected_objects = [dict(row) for row in cursor.fetchall()]
            selected_objects_codes = [row["code"] for row in selected_objects]
            differences = list(required_codes.difference(selected_objects_codes))
            if differences:
                return web.json_response(
                    {
                        "success": False,
                        "reason": f"Не найдены объекты со следующими кодами: {', '.join(map(str, differences))}",
                    },
                    dumps=pretty_json,
                )
            insert_data = []
            for row in table_data:
                if row["VID"] not in [1, 2, 3]:
                    continue
                required_object_id = next(
                    (
                        item["id"]
                        for item in selected_objects
                        if item["code"] == row["KO"]
                    ),
                    None,
                )
                if not required_object_id:
                    raise
                insert_data.append(
                    {
                        "operation_month": app_info["month"],
                        "operation_year": app_info["year"],
                        "payment_month": app_info["month"],
                        "payment_year": app_info["year"],
                        "object_id": required_object_id,
                        "payment_type": row["VID"],
                        "ncen": row["NCEN"],
                        "applied_rate_value": row["TARIF"],
                        "heating_value": row["OTG"],
                        "heating_cost": round(row["TARIF"] * row["OTG"], 5),
                        "water_heating_value": row["GVG"],
                        "water_heating_cost": round(row["TARIF"] * row["GVG"], 5),
                    }
                )
            await conn.execute(self.model.insert(), insert_data)
        return web.json_response(
            {"success": True, "items": insert_data}, dumps=pretty_json
        )


class ObjectPaymentListView(ListView):
    model = payments

    async def get(self):
        object_id = int(self.request.match_info["id"])
        async with self.request.app["db"].connect() as conn:
            cursor = await conn.execute(
                select(self.model)
                .where(self.model.c.object_id == object_id)
                .order_by(
                    desc(self.model.c.operation_month), self.model.c.operation_year
                )
            )
            result = [dict(row) for row in cursor.fetchall()]
            return web.json_response({"success": True, "items": result})

    async def post(self):
        app_info = self.request.app["subsystem"]
        post_data = await self.request.json()
        async with self.request.app["db"].begin() as conn:
            cursor = await conn.execute(
                self.model.insert()
                .returning(literal_column("*"))
                .values(
                    operation_month=app_info["month"],
                    operation_year=app_info["year"],
                    **post_data,
                )
            )
            result = dict(cursor.fetchone())
        return web.json_response({"success": True, "item": result})


class ObjectPaymentsDetailView(DetailView):
    model = payments

    async def patch(self):
        obj_id = int(self.request.match_info["obj_id"])
        payment_id = int(self.request.match_info["id"])
        post_data = await self.request.json()
        async with self.request.app["db"].begin() as conn:
            cursor = await conn.execute(
                self.model.update()
                .returning(literal_column("*"))
                .values(**post_data)
                .where(
                    and_(
                        self.model.c.id == payment_id, self.model.c.object_id == obj_id
                    )
                )
            )
            result = dict(cursor.fetchone())
            return web.json_response(
                {"success": True, "item": result}, dumps=pretty_json
            )

    async def delete(self):
        obj_id = int(self.request.match_info["obj_id"])
        payment_id = int(self.request.match_info["id"])
        async with self.request.app["db"].begin() as conn:
            await conn.execute(
                self.model.delete().where(
                    and_(
                        self.model.c.id == payment_id, self.model.c.object_id == obj_id
                    )
                )
            )
            return web.json_response({"success": True}, dumps=pretty_json)


class RenterPaymentListView(ListView):
    model = payments

    async def get(self):
        renter_id = int(self.request.match_info["id"])
        async with self.request.app["db"].connect() as conn:
            renter_payments = await get_renters_payments(conn, renter_ids=renter_id)
            calculations = await get_renter_payments_calculation(renter_payments)
            return web.json_response(
                {"success": True, "items": calculations}, dumps=pretty_json
            )


class WorkshopsGroupsListView(ListView):
    model = workshops_groups


class WorkshopsGroupDetailView(DetailView):
    model = workshops_groups


class ReviseListView(BaseView):
    async def get(self):
        month = self.request.app["subsystem"]["month"]
        year = self.request.app["subsystem"]["year"]
        async with self.request.app["db"].connect() as conn:
            codes = await get_reconciliation_codes_payments(conn, month, year)
            calculations = await get_reconciliation_codes_payments_calculation(codes)
            return web.json_response({"success": True, "items": calculations})


class FileReportsView(BaseView):
    async def get(self):
        report_name = self.request.match_info["name"]
        month = self.request.app["subsystem"]["month"]
        year = self.request.app["subsystem"]["year"]
        async with self.request.app["db"].connect() as conn:
            if report_name == "consolidated":
                workshop_groups_payments = await get_workshop_groups_payments(
                    conn, month, year
                )
                current_currency_coefficient = await get_current_currency_coefficient(
                    conn, month, year
                )
                calculations = await get_workshops_calculation(workshop_groups_payments)
                return await build_consolidated_report(
                    calculations, month, year, current_currency_coefficient
                )
            elif report_name == "workshop":
                workshop_id = int(self.request.query.get("id", None))
                if not workshop_id:
                    return web.json_response(
                        {
                            "success": False,
                            "reason": "Не верно указан идентификатор цеха",
                        }
                    )
                workshop_payments = await get_workshop_payments(
                    conn, workshop_id, month, year
                )
                calculations = await get_workshop_objects_calculation(workshop_payments)
                return await build_workshop_report(calculations, month, year)
            elif report_name == "renter_short":
                renters_payments = await get_renters_payments(
                    conn, month=month, year=year
                )
                calculations = await get_renter_payments_calculation_short(
                    renters_payments
                )
                return await build_renter_short_report(calculations, month, year)
            elif report_name == "renter_full":
                renters_payments = await get_renters_payments(
                    conn, month=month, year=year
                )
                return await build_renter_full_report(renters_payments, month, year)
            elif report_name == "renter_bank":
                renters_payments = await get_renters_payments(
                    conn, month=month, year=year, is_bank_payment=True
                )
                renters_detailed_payments = await get_renter_detailed_calculation(
                    renters_payments
                )
                return await build_renter_bank_report(
                    renters_detailed_payments, month, year
                )
            elif report_name == "renter_invoice":
                request_items = dict(self.request.query.items())
                required_renters = request_items.get("renters")
                if required_renters is None:
                    return web.json_response(
                        {
                            "success": False,
                            "reason": "Список арендаторов не может быть пустым",
                        }
                    )
                renters_ids = [row["id"] for row in json.loads(required_renters)]
                renter_payments = await get_renters_payments(
                    conn, renter_ids=renters_ids, month=month, year=year
                )
                calculation = await get_renter_vat_calculations(renter_payments)
                return await build_renters_invoices_report(calculation, month, year)
            elif report_name == "renter_invoice_print":
                request_items = dict(self.request.query.items())
                required_renters = request_items.get("renters")
                if required_renters is None:
                    return web.json_response(
                        {
                            "success": False,
                            "reason": "Список арендаторов не может быть пустым",
                        }
                    )
                required_renters_details = json.loads(required_renters)
                renters_ids = [row["id"] for row in required_renters_details]
                renter_payments = await get_renters_payments(
                    conn, renter_ids=renters_ids, month=month, year=year
                )
                calculation = await get_renter_vat_calculations(renter_payments)
                return await build_renters_invoices_print_report(
                    calculation, month, year, required_renters_details
                )
            else:
                return web.json_response(
                    {"success": False, "reason": "Отчет не найден"}
                )

    async def post(self):
        report_name = self.request.match_info["name"]
        month = self.request.app["subsystem"]["month"]
        year = self.request.app["subsystem"]["year"]

        if report_name == "rsc_payments":
            post_data = await self.request.post()
            file = post_data["file"].file
            if not file:
                return web.json_response({"success": False, "reason": "Не указан фаил"})
            workbook = load_workbook(file)
            work_list = workbook[workbook.sheetnames[0]]
            objects_codes = [
                row[0] for row in list(work_list.values)[4:] if isinstance(row[0], int)
            ]
            async with self.request.app["db"].begin() as conn:
                objects_payments = await get_communal_objects_payments(
                    conn, month=month, year=year
                )
            calculated_data = await calculate_communal_objects_values(
                objects_codes, objects_payments
            )
            for index, line in enumerate(work_list):
                if index < 4:
                    continue
                line_code = line[0].value
                if not line_code:
                    break
                payment_values = calculated_data.get(str(line_code))
                if payment_values is None:
                    line[5].value, line[6].value = "", ""
                    continue
                line[5].value = payment_values["heating"] or ""
                line[6].value = payment_values["water_heating"] or ""
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            return web.Response(
                body=output.getvalue(),  # Получаем содержимое
                status=200,
                headers={
                    "Content-Disposition": f"attachment; filename=rsc_report.xlsx",
                    "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                },
            )
        else:
            return web.json_response({"success": False, "reason": "Отчет не найден"})
